from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path(r"D:\project\gacha_performance_test\performance_test_business_optimized.xlsx")
OUTPUT_PATH = Path(r"D:\project\gacha_performance_test\performance_test_business_optimized_reverted_06.xlsx")
SHEET_NAME = "06_执行记录"


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    # Undo the constrained section merges and restore the wide original bands.
    merge_map = {
        "A4:U4": "A4:XFD4",
        "A22:U22": "A22:XFD22",
        "A35:U35": "A35:XFD35",
        "A44:U44": "A44:XFC44",
    }
    current_merges = {str(rng) for rng in ws.merged_cells.ranges}
    for current_range, target_range in merge_map.items():
        if current_range in current_merges:
            ws.unmerge_cells(current_range)
            ws.merge_cells(target_range)

    # Restore the original visible working columns on the left side.
    widths = {
        "A": 10.0,
        "B": 14.0,
        "C": 21.9083333333333,
        "D": 37.9333333333333,
        "E": 78.675,
        "F": 10.0,
        "G": 13.0,
        "H": 12.0,
        "I": 13.0,
        "J": 24.375,
        "K": 10.0,
        "L": 13.0,
        "M": 12.0,
        "N": 13.0,
        "O": 13.0,
        "P": 13.0,
        "Q": 13.0,
        "R": 13.0,
        "S": 13.0,
        "T": 13.0,
        "U": 14.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Screenshot area must stay visible.
    for col_idx in range(22, 64):  # V:BK
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].hidden = False
    # Remove our explicit hide on the following columns as well.
    for col_idx in range(64, 120):  # BL onward
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].hidden = False

    # Restore pre-optimization view settings.
    ws.freeze_panes = "A45"
    ws.sheet_view.zoomScale = 85
    ws.print_area = None
    ws.print_title_rows = None

    # Remove the conditional formatting added during optimization.
    ws.conditional_formatting._cf_rules.clear()

    # Restore the main title/header sizes observed before optimization.
    ws.row_dimensions[1].height = 24.0
    ws.row_dimensions[3].height = 40.0

    # Restore section title presentation.
    yellow_fill = PatternFill("solid", fgColor="FFFFFF00")
    blue_fill = PatternFill("solid", fgColor="FF00B0F0")

    for cell_ref in ("A4", "A22", "A35"):
        cell = ws[cell_ref]
        cell.fill = yellow_fill
        cell.font = Font(name="Microsoft YaHei", size=36, bold=(cell_ref == "A4"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    cell = ws["A44"]
    cell.fill = blue_fill
    cell.font = Font(name="Microsoft YaHei", size=36, bold=False)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Restore key row heights captured before optimization.
    ws.row_dimensions[4].height = 66.0
    ws.row_dimensions[5].height = 207.0
    ws.row_dimensions[8].height = 287.0
    ws.row_dimensions[22].height = 75.0
    ws.row_dimensions[35].height = 89.0
    ws.row_dimensions[44].height = 128.0
    ws.row_dimensions[45].height = 167.0
    ws.row_dimensions[49].height = 177.0

    # Set other data rows back to moderate heights close to the original layout.
    for row_idx in range(6, 22):
        if row_idx != 8:
            ws.row_dimensions[row_idx].height = 207.0
    for row_idx in range(23, 35):
        ws.row_dimensions[row_idx].height = 207.0
    for row_idx in range(36, 44):
        ws.row_dimensions[row_idx].height = 207.0
    for row_idx in range(46, 49):
        ws.row_dimensions[row_idx].height = 167.0
    for row_idx in range(50, 53):
        ws.row_dimensions[row_idx].height = 177.0

    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()
